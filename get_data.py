import serial
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime

# Configurar a porta serial
ser = serial.Serial('/dev/ttyUSB0', 115200)

# Criar workbook e planilha
wb = Workbook()
ws = wb.active
ws.title = "Vazao"

# Cabeçalhos
headers = ["Timestamp", "Pulsos", "Vazão (L/min)", "Vazão (L/s)"]
ws.append(headers)

# Formatar cabeçalhos (negrito e centralizado)
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Ajustar largura das colunas
col_widths = [20, 12, 15, 15]
for i, width in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Espaçamento entre as linhas (altura maior)
ws.row_dimensions[1].height = 25  # cabeçalho
for i in range(2, 1000):
    ws.row_dimensions[i].height = 20  # linhas de dados

try:
    while True:
        linha = ser.readline().decode(errors="ignore").strip()
        if "Pulsos" in linha:  # filtra só as linhas do sensor
            try:
                # Espera formato: "Pulsos: 42 | Vazão: 1.23 L/min | 0.0205 L/s"
                partes = linha.replace("Pulsos:", "").replace("Vazão:", "").replace("L/min", "").replace("L/s", "").split("|")

                pulsos = int(partes[0].strip())
                vazao_l_min = float(partes[1].strip())
                vazao_l_s = float(partes[2].strip())
                ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                # Adicionar linha à planilha
                ws.append([ts, pulsos, vazao_l_min, vazao_l_s])

                # Centralizar novas células
                for col in range(1, 5):
                    ws.cell(row=ws.max_row, column=col).alignment = Alignment(horizontal="center", vertical="center")

                print(ts, pulsos, vazao_l_min, vazao_l_s)

            except Exception:
                print("Linha ignorada:", linha)

except KeyboardInterrupt:
    # Salvar arquivo ao parar o script
    wb.save("dados_vazao.xlsx")
    print("Planilha salva como dados_vazao.xlsx")
    ser.close()
